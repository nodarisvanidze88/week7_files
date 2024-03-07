import openpyxl

data = [
        ["Apple", 1.99, 100],
        ["Banana", 0.99, 150],
        ["Orange", 1.49, 120],
        ["watermalon", 5.2, 20]
    ]

def create_excel_file(ex_file):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"

    headers = ["Product Name", "Price", "Quantity"]
    sheet.append(headers)

    for item in data:
        sheet.append(item)
    workbook.save(ex_file)

def read_excel_file(ex_file):
    workbook = openpyxl.load_workbook(ex_file)
    sheet = workbook.active

    for row in sheet.iter_rows(values_only=True):
        print(row)

        
def modify_excel_file(ex_file, row, col, new_val):
    workbook = openpyxl.load_workbook(ex_file)
    sheet = workbook.active
    sheet.cell(row=row,column=col,value = new_val)
    workbook.save(ex_file)
# create_excel_file("text_excel.xlsx")
# read_excel_file("text_excel.xlsx")
modify_excel_file("text_excel.xlsx",2,2,"Welcome")
